import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# 共用工具
def get_fill_color(cell):
    fill = cell.fill

    if not fill:
        return None
KH_ROW_LABEL_RULES
    color = fill.fgColor

    if not color:
        return None

    if color.type == "rgb":
        return color.rgb

    # theme 色先忽略 tint，直接記 theme index
    if color.type == "theme":
        tint = getattr(color, "tint", 0)
        return f"theme:{color.theme}:tint:{tint}"

    if color.type == "indexed":
        return f"indexed:{color.indexed}"

    return None


def normalize_row_label(value):
    if value is None:
        return ""

    if isinstance(value, (int, float)):
        if int(value) == value:
            return str(int(value))
        return str(value)

    return str(value).strip()


def label_to_zone_price_available(label):
    text = str(label).strip()

    # 特殊席
    if "工作席" in text:
        return "staff", 0, False

    if "攝影" in text:
        return "camera", 0, False

    if "貴賓" in text:
        return "vip", 0, False

    if "輪椅陪同" in text:
        return "companion", 300, False

    if "輪椅" in text:
        return "wheelchair", 300, False

    if "不開放" in text:
        return "notopen", 300, False

    # 團內票區
    if "團內" in text:
        if "800" in text:
            return "group-800", 640, True
        if "500" in text:
            return "group-500", 400, True
        if "300" in text:
            return "group-300", 240, True
        if "200" in text:
            return "group-200", 160, True

    # 一般票區
    if "800" in text:
        return "regular-800", 800, False
    if "500" in text:
        return "regular-500", 500, False
    if "300" in text:
        return "regular-300", 300, False
    if "200" in text:
        return "regular-200", 200, False

    return "unknown", 0, False


# =========================
# 主入口
# =========================

def parse_seat_map(filepath, concert_code="tp", debug=False):
    if concert_code == "tp":
        return parse_tp_seat_map(filepath, debug=debug)

    if concert_code == "kh":
        return parse_kh_seat_map(filepath, debug=debug)

    raise ValueError(f"未知場次：{concert_code}")


# =========================
# 台北場 parser
# =========================

TP_SEAT_START_COL = column_index_from_string("E")
TP_SEAT_END_COL = column_index_from_string("AT")

TP_LEFT_ROW_LABEL_COL = column_index_from_string("B")
TP_RIGHT_ROW_LABEL_COL = column_index_from_string("AW")

TP_LEGEND_COLOR_COL = column_index_from_string("AX")
TP_LEGEND_LABEL_COL = column_index_from_string("AY")


def build_tp_legend_label_map(ws):
    legend_map = {}

    for row in range(1, ws.max_row + 1):
        color_cell = ws.cell(row, TP_LEGEND_COLOR_COL)
        label_cell = ws.cell(row, TP_LEGEND_LABEL_COL)

        color = get_fill_color(color_cell)
        label = label_cell.value

        if not color or label is None:
            continue

        label = str(label).strip()
        if not label:
            continue

        legend_map[color] = label

    return legend_map


def build_tp_color_map(ws):
    legend_label_map = build_tp_legend_label_map(ws)
    color_map = {}

    for color, label in legend_label_map.items():
        zone, price, available = label_to_zone_price_available(label)
        color_map[color] = (zone, price, available)

    return color_map


def parse_tp_seat_map(filepath, debug=False):
    wb = load_workbook(filepath)
    ws = wb.active

    color_map = build_tp_color_map(ws)

    seats = []
    row_labels = {}

    for excel_row in range(1, ws.max_row + 1):
        left_label = normalize_row_label(
            ws.cell(excel_row, TP_LEFT_ROW_LABEL_COL).value
        )
        right_label = normalize_row_label(
            ws.cell(excel_row, TP_RIGHT_ROW_LABEL_COL).value
        )

        row_label = left_label or right_label

        if row_label:
            row_labels[excel_row] = row_label

        for excel_col in range(TP_SEAT_START_COL, TP_SEAT_END_COL + 1):
            cell = ws.cell(excel_row, excel_col)
            value = cell.value

            if isinstance(value, (int, float)):
                color = get_fill_color(cell)
                zone, price, available = color_map.get(
                    color,
                    ("unknown", 0, False)
                )

                if debug and color not in color_map:
                    print(
                        f"⚠️ 台北場未定義顏色: "
                        f"row={excel_row}, col={excel_col}, "
                        f"seat={value}, color={color}"
                    )

                seats.append({
                    "seat_number": int(value),
                    "excel_row": excel_row,
                    "excel_col": excel_col,
                    "row_label": row_labels.get(excel_row, ""),
                    "zone": zone,
                    "price": price,
                    "color": color,
                    "available": available,
                })

    return seats, row_labels, color_map

# =========================
# 高雄場 parser
# =========================

KH_SCAN_START_COL = column_index_from_string("C")
KH_SCAN_END_COL = column_index_from_string("CY")

KH_SCAN_START_ROW = 6
KH_SCAN_END_ROW = 88

KH_STAGE_START_COL = column_index_from_string("AJ")
KH_STAGE_END_COL = column_index_from_string("BV")
KH_STAGE_START_ROW = 35
KH_STAGE_END_ROW = 41

KH_FIRST_FLOOR_START_COL = column_index_from_string("AJ")
KH_FIRST_FLOOR_END_COL = column_index_from_string("BU")
KH_FIRST_FLOOR_START_ROW = 44
KH_FIRST_FLOOR_END_ROW = 55

KH_SECOND_FLOOR_START_COL = column_index_from_string("R")
KH_SECOND_FLOOR_END_COL = column_index_from_string("CM")
KH_SECOND_FLOOR_START_ROW = 17
KH_SECOND_FLOOR_END_ROW = 82

# 高雄場 legend：目前確認是 DB / DC
KH_LEGEND_COLOR_COL = column_index_from_string("DB")
KH_LEGEND_LABEL_COL = column_index_from_string("DC")
KH_LEGEND_START_ROW = 30
KH_LEGEND_END_ROW = 45

# 高雄場排數欄位，不是座位
KH_ROW_NUMBER_COLS = {
    column_index_from_string("AI"),
    column_index_from_string("AT"),
    column_index_from_string("AV"),
    column_index_from_string("BI"),
    column_index_from_string("BK"),
    column_index_from_string("BV"),
}

# 一樓排數欄位
KH_FIRST_FLOOR_ROW_MARKER_COLS = {
    column_index_from_string("AI"),
    column_index_from_string("AT"),
    column_index_from_string("AV"),
    column_index_from_string("BI"),
    column_index_from_string("BK"),
}

# 高雄場手動排數規則
# mode = "col"：直排，依照欄位 + row range 判斷
# mode = "row"：橫排，依照列 + col range 判斷
KH_ROW_LABEL_RULES = [
    # ===== 二樓：左側直排 =====
    {"label": "E3", "mode": "col", "col": "T", "start_row": 42, "end_row": 50},
    {"label": "E2", "mode": "col", "col": "U", "start_row": 35, "end_row": 52},
    {"label": "E1", "mode": "col", "col": "V", "start_row": 25, "end_row": 57},
    {"label": "C1", "mode": "col", "col": "V", "start_row": 61, "end_row": 77},
    {"label": "C2", "mode": "col", "col": "T", "start_row": 75, "end_row": 78},

    {"label": "D4", "mode": "col", "col": "Y", "start_row": 33, "end_row": 55},
    {"label": "D3", "mode": "col", "col": "Z", "start_row": 31, "end_row": 55},
    {"label": "D2", "mode": "col", "col": "AB", "start_row": 34, "end_row": 54},
    {"label": "D1", "mode": "col", "col": "AC", "start_row": 37, "end_row": 53},

    # ===== 二樓：右側直排 =====
    {"label": "D1", "mode": "col", "col": "CB", "start_row": 37, "end_row": 53},
    {"label": "D2", "mode": "col", "col": "CC", "start_row": 34, "end_row": 54},
    {"label": "D3", "mode": "col", "col": "CE", "start_row": 31, "end_row": 55},
    {"label": "D4", "mode": "col", "col": "CF", "start_row": 33, "end_row": 55},

    {"label": "E1", "mode": "col", "col": "CI", "start_row": 25, "end_row": 57},
    {"label": "E2", "mode": "col", "col": "CJ", "start_row": 35, "end_row": 52},
    {"label": "E3", "mode": "col", "col": "CK", "start_row": 33, "end_row": 51},
    {"label": "C1", "mode": "col", "col": "CH", "start_row": 61, "end_row": 77},
    {"label": "C2", "mode": "col", "col": "CJ", "start_row": 74, "end_row": 78},

    # ===== 三樓：左側 =====
    {"label": "B4", "mode": "col", "col": "J", "start_row": 43, "end_row": 49},
    {"label": "B3", "mode": "col", "col": "K", "start_row": 42, "end_row": 49},
    {"label": "B2", "mode": "col", "col": "L", "start_row": 41, "end_row": 49},

    {"label": "B4", "mode": "col", "col": "J", "start_row": 58, "end_row": 72},
    {"label": "B3", "mode": "col", "col": "K", "start_row": 53, "end_row": 72},
    {"label": "B2", "mode": "col", "col": "L", "start_row": 53, "end_row": 72},
    {"label": "B1", "mode": "col", "col": "M", "start_row": 57, "end_row": 67},

    # ===== 三樓：右側 =====
    {"label": "B2", "mode": "col", "col": "CR", "start_row": 31, "end_row": 51},
    {"label": "B3", "mode": "col", "col": "CS", "start_row": 31, "end_row": 51},
    {"label": "B4", "mode": "col", "col": "CT", "start_row": 31, "end_row": 51},

    {"label": "B1", "mode": "col", "col": "CQ", "start_row": 59, "end_row": 69},
    {"label": "B2", "mode": "col", "col": "CR", "start_row": 55, "end_row": 74},
    {"label": "B3", "mode": "col", "col": "CS", "start_row": 55, "end_row": 74},
    {"label": "B4", "mode": "col", "col": "CT", "start_row": 59, "end_row": 73},

    # ===== 二樓：A 區橫排 =====
    {"label": "A1", "mode": "row", "row": 61, "start_col": "AC", "end_col": "CA"},
    {"label": "A2", "mode": "row", "row": 62, "start_col": "AC", "end_col": "CA"},
    {"label": "A3", "mode": "row", "row": 63, "start_col": "AC", "end_col": "CA"},
    {"label": "A4", "mode": "row", "row": 64, "start_col": "AC", "end_col": "CA"},
    {"label": "A5", "mode": "row", "row": 65, "start_col": "AC", "end_col": "CA"},

    # A6：左右兩段
    {"label": "A6", "mode": "row", "row": 66, "start_col": "AH", "end_col": "AM"},
    {"label": "A6", "mode": "row", "row": 66, "start_col": "BR", "end_col": "BW"},

    # ===== 二樓：B 區橫排 =====
    {"label": "B1", "mode": "row", "row": 66, "start_col": "AW", "end_col": "BG"},
    {"label": "B2", "mode": "row", "row": 67, "start_col": "AR", "end_col": "BM"},

    {"label": "B3", "mode": "row", "row": 69, "start_col": "AO", "end_col": "BP"},
    {"label": "B4", "mode": "row", "row": 70, "start_col": "AO", "end_col": "BP"},
    {"label": "B5", "mode": "row", "row": 71, "start_col": "AO", "end_col": "BP"},
    {"label": "B6", "mode": "row", "row": 72, "start_col": "AO", "end_col": "BP"},
    {"label": "B7", "mode": "row", "row": 73, "start_col": "AO", "end_col": "BP"},
    {"label": "B8", "mode": "row", "row": 74, "start_col": "AR", "end_col": "BM"},
    {"label": "B9", "mode": "row", "row": 75, "start_col": "AR", "end_col": "BM"},

    # ===== 二樓：C 區 L-shape 橫排 =====
    {"label": "C1", "mode": "row", "row": 78, "start_col": "AS", "end_col": "BK"},
    {"label": "C2", "mode": "row", "row": 79, "start_col": "AR", "end_col": "BM"},
    {"label": "C3", "mode": "row", "row": 80, "start_col": "Q", "end_col": "CN"},
    {"label": "C4", "mode": "row", "row": 81, "start_col": "U", "end_col": "CK"},
    {"label": "C5", "mode": "row", "row": 82, "start_col": "W", "end_col": "CH"},
]


def is_in_range(row, col, start_row, end_row, start_col, end_col):
    return (
        start_row <= row <= end_row
        and start_col <= col <= end_col
    )


def col_to_index(col_name):
    return column_index_from_string(col_name)


def is_kh_stage(row, col):
    return is_in_range(
        row,
        col,
        KH_STAGE_START_ROW,
        KH_STAGE_END_ROW,
        KH_STAGE_START_COL,
        KH_STAGE_END_COL,
    )


def get_kh_floor(row, col):
    if is_in_range(
        row,
        col,
        KH_FIRST_FLOOR_START_ROW,
        KH_FIRST_FLOOR_END_ROW,
        KH_FIRST_FLOOR_START_COL,
        KH_FIRST_FLOOR_END_COL,
    ):
        return "1樓"

    if is_in_range(
        row,
        col,
        KH_SECOND_FLOOR_START_ROW,
        KH_SECOND_FLOOR_END_ROW,
        KH_SECOND_FLOOR_START_COL,
        KH_SECOND_FLOOR_END_COL,
    ):
        return "2樓"

    return "3樓"


def get_kh_row_label(row, col):
    # 一樓：44 = 1排，55 = 12排
    if is_in_range(
        row,
        col,
        KH_FIRST_FLOOR_START_ROW,
        KH_FIRST_FLOOR_END_ROW,
        KH_FIRST_FLOOR_START_COL,
        KH_FIRST_FLOOR_END_COL,
    ):
        return str(row - KH_FIRST_FLOOR_START_ROW + 1)

    # 二樓 / 三樓：依照人工規則判斷
    for rule in KH_ROW_LABEL_RULES:
        if rule["mode"] == "col":
            rule_col = col_to_index(rule["col"])

            if (
                col == rule_col
                and rule["start_row"] <= row <= rule["end_row"]
            ):
                return rule["label"]

        elif rule["mode"] == "row":
            start_col = col_to_index(rule["start_col"])
            end_col = col_to_index(rule["end_col"])

            if (
                row == rule["row"]
                and start_col <= col <= end_col
            ):
                return rule["label"]

    return ""


def build_kh_color_map(ws):
    color_map = {}

    for row in range(KH_LEGEND_START_ROW, KH_LEGEND_END_ROW + 1):
        color_cell = ws.cell(row, KH_LEGEND_COLOR_COL)
        label_cell = ws.cell(row, KH_LEGEND_LABEL_COL)

        color = get_fill_color(color_cell)
        label = label_cell.value

        if not color or color == "00000000" or label is None:
            continue

        label = str(label).strip()
        if not label:
            continue

        zone, price, available = label_to_zone_price_available(label)
        color_map[color] = (zone, price, available)

    color_map["theme:8:tint:-0.499984740745262"] = ("notopen", 300, False)

    return color_map


def build_kh_merged_lookup(ws):
    merged_lookup = {}

    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        if (
            max_row < KH_SCAN_START_ROW
            or min_row > KH_SCAN_END_ROW
            or max_col < KH_SCAN_START_COL
            or min_col > KH_SCAN_END_COL
        ):
            continue

        row_span = max_row - min_row + 1
        col_span = max_col - min_col + 1

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_lookup[(row, col)] = {
                    "min_row": min_row,
                    "max_row": max_row,
                    "min_col": min_col,
                    "max_col": max_col,
                    "row_span": row_span,
                    "col_span": col_span,
                    "is_top_left": row == min_row and col == min_col,
                }

    return merged_lookup


def parse_kh_seat_map(filepath, debug=False):
    wb = load_workbook(filepath)
    ws = wb.active

    color_map = build_kh_color_map(ws)
    merged_lookup = build_kh_merged_lookup(ws)

    seats = []
    row_labels = {}

    for excel_row in range(KH_SCAN_START_ROW, KH_SCAN_END_ROW + 1):
        for excel_col in range(KH_SCAN_START_COL, KH_SCAN_END_COL + 1):

            if is_kh_stage(excel_row, excel_col):
                continue

            if excel_col in KH_ROW_NUMBER_COLS:
                continue

            cell = ws.cell(excel_row, excel_col)
            value = cell.value

            if not isinstance(value, (int, float)):
                continue

            color = get_fill_color(cell)

            if not color or color == "00000000":
                continue

            merged_info = merged_lookup.get((excel_row, excel_col))

            if merged_info:
                if not merged_info["is_top_left"]:
                    continue

                is_four_cell_merge = (
                    merged_info["row_span"] == 2
                    and merged_info["col_span"] == 2
                )

                zone, price, available = color_map.get(
                    color,
                    ("unknown", 0, False)
                )

                if is_four_cell_merge:
                    if zone == "wheelchair":
                        available = False
                    else:
                        if debug:
                            print(
                                f"🚪 高雄場 gate 已略過: "
                                f"row={excel_row}, col={excel_col}, "
                                f"value={value}, color={color}"
                            )
                        continue

            else:
                zone, price, available = color_map.get(
                    color,
                    ("unknown", 0, False)
                )

            if zone in ("wheelchair", "companion"):
                available = False

            floor = get_kh_floor(excel_row, excel_col)
            row_label = get_kh_row_label(excel_row, excel_col)

            if row_label:
                row_labels[excel_row] = row_label

            if debug and color not in color_map:
                print(
                    f"⚠️ 高雄場未定義顏色: "
                    f"row={excel_row}, col={excel_col}, "
                    f"seat={value}, color={color}"
                )

            seats.append({
                "seat_number": int(value),
                "excel_row": excel_row,
                "excel_col": excel_col,
                "row_label": row_label,
                "floor": floor,
                "zone": zone,
                "price": price,
                "color": color,
                "available": available,
            })

    return seats, row_labels, color_map
